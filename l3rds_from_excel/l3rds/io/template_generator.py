"""Excel template generator for lower thirds data entry.

This module creates professionally formatted Excel templates with three sheets:
1. Data - Empty sheet with headers for user input
2. Examples - Sample rows demonstrating features
3. Instructions - Comprehensive guide for filling out the template
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from l3rds.utils.logger import get_logger

logger = get_logger(__name__)


class TemplateGenerator:
    """Generates Excel templates for lower thirds data entry."""

    # All columns in order
    COLUMNS = [
        "Main Text",
        "Secondary Text",
        "Justification",
        "Main Font",
        "Secondary Font",
        "Main Font Size",
        "Secondary Font Size",
        "Main Color",
        "Secondary Color",
        "Background Color",
        "Bar Color",
        "Text Outline",
        "Text Shadow",
        "Shadow Color",
        "Padding",
        "Position Offset X",
        "Position Offset Y",
        "Wrap Text",
        "Wrap Padding",
        "File Name",
    ]

    # Example data rows
    EXAMPLES = [
        # Minimal example
        {
            "Main Text": "John Smith",
            "Secondary Text": "Director",
            "Justification": "lower left",
        },
        # Basic customization
        {
            "Main Text": "Jane Doe",
            "Secondary Text": "Senior Engineer",
            "Justification": "lower center",
            "Main Font": "Arial",
            "Main Color": "white",
            "Secondary Color": "light blue",
        },
        # Advanced effects
        {
            "Main Text": "Bob Wilson",
            "Secondary Text": "CEO",
            "Justification": "upper right",
            "Main Font": "Helvetica",
            "Main Color": "yellow",
            "Bar Color": "blue,128",
            "Text Shadow": "Yes",
            "Shadow Color": "black",
            "Text Outline": "2,black,255",
        },
        # Custom layout
        {
            "Main Text": "Sarah Johnson",
            "Secondary Text": "Marketing Manager",
            "Justification": "center center",
            "Main Font Size": "72",
            "Secondary Font Size": "36",
            "Padding": "50",
            "Position Offset X": "10",
            "Position Offset Y": "-20",
            "Wrap Text": "Yes",
            "Wrap Padding": "100",
        },
        # Custom filename
        {
            "Main Text": "Michael Chen",
            "Secondary Text": "Creative Director",
            "Justification": "lower right",
            "File Name": "mike_chen_lower_third",
        },
    ]

    @classmethod
    def create_template(cls, output_path: str | Path) -> None:
        """Create Excel template with data, examples, and instructions sheets.

        Args:
            output_path: Path where template will be saved

        Raises:
            Exception: If template creation fails
        """
        logger.info(f"Creating Excel template: {output_path}")

        try:
            # Create workbook
            wb = Workbook()

            # Create Data sheet (rename default active sheet)
            ws_data = wb.active
            ws_data.title = "Data"
            cls._create_data_sheet(ws_data)

            # Create Examples sheet
            ws_examples = wb.create_sheet("Examples")
            cls._create_examples_sheet(ws_examples)

            # Create Instructions sheet
            ws_instructions = wb.create_sheet("Instructions")
            cls._create_instructions_sheet(ws_instructions)

            # Save workbook
            wb.save(output_path)
            logger.info(f"Excel template created successfully: {output_path}")

        except Exception as e:
            logger.error(f"Failed to create Excel template: {e}")
            raise

    @classmethod
    def _create_data_sheet(cls, ws) -> None:
        """Create the main data entry sheet with headers.

        Args:
            ws: Worksheet to populate
        """
        # Write headers with formatting
        for col_idx, col_name in enumerate(cls.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cls._format_header_cell(cell)

        # Auto-adjust column widths
        for col_idx in range(1, len(cls.COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        # Freeze header row
        ws.freeze_panes = "A2"

    @classmethod
    def _create_examples_sheet(cls, ws) -> None:
        """Create examples sheet with sample data rows.

        Args:
            ws: Worksheet to populate
        """
        # Write headers
        for col_idx, col_name in enumerate(cls.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cls._format_header_cell(cell)

        # Write example rows
        for row_idx, example in enumerate(cls.EXAMPLES, start=2):
            for col_idx, col_name in enumerate(cls.COLUMNS, start=1):
                value = example.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Light formatting for example rows
                cell.alignment = Alignment(horizontal='left', vertical='center')

        # Auto-adjust column widths
        for col_idx in range(1, len(cls.COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        # Freeze header row
        ws.freeze_panes = "A2"

    @classmethod
    def _create_instructions_sheet(cls, ws) -> None:
        """Create instructions sheet with comprehensive guide.

        Args:
            ws: Worksheet to populate
        """
        # Set column width for readability
        ws.column_dimensions['A'].width = 120

        row = 1

        # Title
        title_cell = ws.cell(row=row, column=1, value="LOWER THIRDS GENERATOR - EXCEL TEMPLATE INSTRUCTIONS")
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 30
        row += 2

        # Introduction
        cls._add_instruction_section(ws, row, "OVERVIEW", [
            "This template helps you create lower third graphics for video production.",
            "Fill in the 'Data' sheet with your text and formatting preferences.",
            "See the 'Examples' sheet for demonstration of features.",
            "",
            "MINIMUM REQUIREMENT: Provide at least one of 'Main Text' or 'Secondary Text' for each row.",
        ])
        row += 7

        # Required columns
        cls._add_instruction_section(ws, row, "REQUIRED COLUMNS (at least one)", [
            "• Main Text - Primary text to display (typically a name)",
            "• Secondary Text - Secondary text to display (typically a title or role)",
            "",
            "Note: At least ONE of these must be filled in for each row.",
        ])
        row += 6

        # Optional columns - Position
        cls._add_instruction_section(ws, row, "POSITION", [
            "• Justification - Text position on screen. Valid values:",
            "  - lower left (traditional lower third)",
            "  - lower center, center bottom",
            "  - lower right",
            "  - upper left",
            "  - upper center, center top",
            "  - upper right",
            "  - center center, center",
            "  Default: lower left",
        ])
        row += 11

        # Fonts
        cls._add_instruction_section(ws, row, "FONTS", [
            "• Main Font - Font name (e.g., 'Arial', 'Helvetica') or path to .ttf/.otf file",
            "  Leave blank to use GUI/config default",
            "• Secondary Font - Font for secondary text. Leave blank to use main font",
            "• Main Font Size - Size in pixels. Leave blank for auto-calculation (height/18)",
            "• Secondary Font Size - Size in pixels. Leave blank for auto-calculation (height/25)",
        ])
        row += 7

        # Colors
        cls._add_instruction_section(ws, row, "COLORS", [
            "All color fields are optional. Leave blank to use GUI/config defaults.",
            "",
            "• Main Color - Color for main text",
            "• Secondary Color - Color for secondary text",
            "• Background Color - Background color for entire image",
            "• Bar Color - Color for lower third bar (can include opacity: 'blue,128')",
            "",
            "COLOR FORMATS:",
            "  - Named: 'red', 'white', 'dark blue', 'sky blue'",
            "  - Hex: '#FF0000' or 'FF0000'",
            "  - RGB: '255,0,0' or 'rgb(255,0,0)'",
            "  - RGBA: 'red,128' or '255,0,0,128' (alpha: 0=transparent, 255=opaque)",
            "",
            "SUPPORTED COLOR NAMES:",
            "black, white, red, green, blue, yellow, cyan, magenta, purple, orange, pink,",
            "gray, grey, brown, navy, teal, lime, maroon, olive, silver, gold, indigo,",
            "violet, turquoise, tan, salmon, sky blue, khaki, crimson, dark blue,",
            "dark green, dark red, dark gray, dark grey, light gray, light grey,",
            "light blue, light green, light red, transparent",
        ])
        row += 24

        # Effects
        cls._add_instruction_section(ws, row, "TEXT EFFECTS", [
            "• Text Outline - Format: 'WIDTH,COLOR,OPACITY'",
            "  Example: '2,black,255' (2 pixel wide black outline, fully opaque)",
            "  Example: '3,white,200' (3 pixel white outline, slightly transparent)",
            "",
            "• Text Shadow - Enable shadow effect",
            "  Values: Yes/No, True/False, 1/0, On/Off, Enabled/Disabled (case-insensitive)",
            "",
            "• Shadow Color - Shadow color specification (any color format)",
        ])
        row += 10

        # Layout
        cls._add_instruction_section(ws, row, "LAYOUT ADJUSTMENTS", [
            "• Padding - Distance from image edge in pixels",
            "• Position Offset X - Horizontal fine-tune adjustment (negative=left, positive=right)",
            "  Range: -500 to 500",
            "• Position Offset Y - Vertical fine-tune adjustment (negative=up, positive=down)",
            "  Range: -500 to 500",
            "",
            "• Wrap Text - Enable text wrapping",
            "  Values: Yes/No, True/False, 1/0",
            "• Wrap Padding - Distance from edge where text wraps (REQUIRED if Wrap Text is Yes)",
        ])
        row += 12

        # Output
        cls._add_instruction_section(ws, row, "OUTPUT", [
            "• File Name - Custom filename (without extension)",
            "  Leave blank to use Main Text or Secondary Text as filename",
            "  Invalid characters will be removed automatically",
        ])
        row += 5

        # Tips
        cls._add_instruction_section(ws, row, "TIPS & BEST PRACTICES", [
            "✓ Leave optional columns blank to use GUI/configuration defaults",
            "✓ Excel values override GUI settings when specified",
            "✓ Use Examples sheet as a starting point - copy and modify",
            "✓ Test with one row first before creating many entries",
            "✓ Color values are case-insensitive",
            "✓ Boolean fields accept various formats: Yes/No, True/False, 1/0, On/Off",
            "✓ Font names: Use system-installed font names or full paths to font files",
            "✓ For consistent styling, use GUI defaults and only override when needed",
        ])
        row += 11

        # Common mistakes
        cls._add_instruction_section(ws, row, "COMMON MISTAKES TO AVOID", [
            "✗ Forgetting to fill in at least one text field (Main Text or Secondary Text)",
            "✗ Misspelling position values (must match exactly: 'lower left', not 'lowerleft')",
            "✗ Using invalid color names (check supported color list above)",
            "✗ Incorrect outline format (must be 'width,color' or 'width,color,opacity')",
            "✗ Enabling Wrap Text but not specifying Wrap Padding",
            "✗ Using quotes around values unnecessarily (Excel handles this)",
            "✗ File paths with spaces - no special quoting needed in Excel",
        ])

    @classmethod
    def _add_instruction_section(cls, ws, start_row: int, title: str, lines: list[str]) -> None:
        """Add a formatted instruction section to the worksheet.

        Args:
            ws: Worksheet to add section to
            start_row: Starting row number
            title: Section title
            lines: List of instruction lines
        """
        # Title
        title_cell = ws.cell(row=start_row, column=1, value=title)
        title_cell.font = Font(bold=True, size=13, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2E75B5", end_color="2E75B5", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[start_row].height = 20

        # Content lines
        for idx, line in enumerate(lines, start=1):
            cell = ws.cell(row=start_row + idx, column=1, value=line)
            cell.font = Font(size=11)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    @classmethod
    def _format_header_cell(cls, cell) -> None:
        """Apply consistent formatting to header cells.

        Args:
            cell: Cell to format
        """
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
